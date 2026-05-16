"""
Generates the Excel model output.
Ports the existing multi-sheet Excel build into the new architecture.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import yaml
from datetime import date


def tb(c='CCCCCC'):
    s = Side(border_style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(cell, bg='1A1A2E', fg='FFFFFF', sz=10, bold=True, left=False):
    cell.font = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(
        horizontal='left' if left else 'center',
        vertical='center', wrap_text=True)
    cell.border = tb()


def val(cell, bg='FFFFFF', bold=False, color='000000',
        sz=9, ha='center', italic=False, wrap=False):
    cell.font = Font(name='Calibri', bold=bold, color=color, size=sz, italic=italic)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border = tb()


def cw(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w


def rh(ws, row, h):
    ws.row_dimensions[row].height = h


def rbg(rating):
    m = {
        'Safe D':   'C8D9EE', 'Likely D': 'EBF3FB',
        'Lean D':   'E3F2FD', 'Tossup':   'F3E5FF',
        'Lean R':   'FBE9E7', 'Likely R': 'FBF0EB',
        'Safe R':   'EEDCC8',
    }
    return m.get(rating, 'FFFFFF')


def rcol(rating):
    if 'D' in rating: return '1B4F8A'
    if 'R' in rating: return '8B1A1A'
    return '555555'


def build_excel(oop, demo, state_envs, race_results, sim, config, out_path):
    from model.race_model import SENATE_RACES_2026

    wb = openpyxl.Workbook()

    # ── SHEET 0: DASHBOARD ────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = '0. Dashboard'
    ws0.sheet_view.showGridLines = False

    ws0.merge_cells('A1:L1')
    c = ws0['A1']
    c.value = (f"2026 SENATE MODEL v1  |  Run: {date.today()}  |  "
               f"Bottom-up MRP  |  Catalist 2024 SSOT")
    hdr(c, sz=13, left=True)
    rh(ws0, 1, 28)

    # KPI row
    kpis = [
        (1,  3,  'D MAJORITY',    f"{sim['d_majority_prob']:.1%}", '1B4F8A'),
        (4,  6,  'EXP D SEATS',   f"{sim['exp_d_seats']:.1f}",    '2C3E6B'),
        (7,  9,  'GCB EDAY',      f"D{oop['gcb_eday_central']:+.2f}", '1B4F8A'),
        (10, 12, 'OOP SHIFT',     f"+{oop['oop_shift']:.2f}pp",   '2C3E6B'),
    ]
    for cs, ce, label, value, bg in kpis:
        for row in [3, 4, 5, 6]:
            ws0.merge_cells(
                f"{get_column_letter(cs)}{row}:{get_column_letter(ce)}{row}")
            for c_idx in range(cs, ce+1):
                ws0.cell(row=row, column=c_idx).fill = PatternFill('solid', fgColor=bg)
        lc = ws0.cell(row=3, column=cs, value=label)
        lc.font = Font(name='Calibri', color='AAAAAA', size=7, bold=True)
        lc.alignment = Alignment(horizontal='center', vertical='center')
        vc = ws0.cell(row=4, column=cs, value=value)
        vc.font = Font(name='Calibri', bold=True, color='FFFFFF', size=22)
        vc.alignment = Alignment(horizontal='center', vertical='center')
        sc = ws0.cell(row=5, column=cs)
        sc.fill = PatternFill('solid', fgColor='0D1B3E')
    for row, h in zip([3, 4, 5, 6], [12, 26, 8, 6]):
        rh(ws0, row, h)

    # Race ratings table
    ws0.merge_cells('A8:L8')
    hdr(ws0.cell(row=8, column=1, value='SENATE RACE RATINGS — COMPETITIVE RACES'),
        bg='2C3E6B', sz=10, left=True)
    rh(ws0, 8, 20)

    headers = ['State', 'Inc', 'D Candidate', 'R Candidate',
               'GCB Eday', 'Poll Blend', 'N', 'WAR Net',
               'σ', 'Central', 'D Win%', 'Rating']
    for ci, h in enumerate(headers, 1):
        hdr(ws0.cell(row=9, column=ci, value=h), bg='2C3E6B', sz=9)
    rh(ws0, 9, 20)

    # Get race info lookup
    race_info = {abbr: (inc, dc, rc, comp)
                 for abbr, inc, dc, rc, comp in SENATE_RACES_2026}

    comp_results = sorted(
        [r for abbr, r in race_results.items()
         if race_info.get(abbr, (None,None,None,False))[3]],
        key=lambda x: -x['d_prob']
    )

    for ri, r in enumerate(comp_results):
        row = ri + 10
        abbr = r['abbr']
        inc, dc, rc, _ = race_info.get(abbr, ('?', 'TBD', 'TBD', True))
        bg2 = rbg(r['rating'])
        rh(ws0, row, 16)

        row_data = [
            (abbr, 'center', True),
            (inc, 'center', False),
            (dc[:22], 'left', False),
            (rc[:18], 'left', False),
            (r['gcb_eday'], 'center', False),
            (r['poll_blend'] if r['poll_blend'] is not None else 'n/a', 'center', False),
            (r['n_polls'], 'center', False),
            (r['war_net'], 'center', False),
            (r['total_sigma'], 'center', False),
            (r['central'], 'center', True),
            (r['d_prob'], 'center', True),
            (r['rating'], 'center', True),
        ]
        for ci, (v, ha, bold) in enumerate(row_data, 1):
            cell = ws0.cell(row=row, column=ci, value=v)
            val(cell, bg=bg2, bold=bold, sz=9, ha=ha)
            if ci in (5, 6, 8, 10) and isinstance(v, float):
                cell.number_format = '+0.0;-0.0;0.0'
            if ci == 9 and isinstance(v, float):
                cell.number_format = '0.00'
            if ci == 11 and isinstance(v, float):
                cell.number_format = '0.0%'
                cell.font = Font(name='Calibri', bold=True, size=9,
                                 color='1B4F8A' if v >= 0.5 else '8B1A1A')
            if ci == 12:
                cell.font = Font(name='Calibri', bold=True, size=9,
                                 color=rcol(r['rating']))

    col_widths = [5, 4, 22, 18, 9, 9, 5, 8, 7, 9, 9, 11]
    for i, w in enumerate(col_widths, 1):
        cw(ws0, i, w)

    # ── SHEET 1: OOP METHODOLOGY ──────────────────────────────────────────────
    ws1 = wb.create_sheet('1. OOP Shift')
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells('A1:I1')
    hdr(ws1.cell(row=1, column=1,
                 value=f"OOP SHIFT METHODOLOGY  |  Shift: {oop['oop_shift']:+.2f}pp  |  "
                       f"GCB today: D{oop['gcb_today']:+.2f}  |  "
                       f"GCB eday: D{oop['gcb_eday_central']:+.2f}"),
        sz=11, left=True)
    rh(ws1, 1, 24)

    for ci, h in enumerate(['Year', 'OOP Shift', 'Recency Weight',
                             'Weighted Contrib', 'Notes'], 1):
        hdr(ws1.cell(row=2, column=ci, value=h), bg='2C3E6B', sz=9)
    rh(ws1, 2, 20)

    cycle_notes = {
        1994: 'Gingrich/Contract with America',
        1998: 'Clinton impeachment backlash — OOP lost ground',
        2006: 'Iraq/Katrina',
        2010: 'Tea Party wave',
        2014: 'Obama 2nd term drag',
        2018: 'Anti-Trump',
        2022: 'Dobbs softened wave',
    }
    for ri, (yr, wt, sh) in enumerate(oop['cycle_weights']):
        r = ri + 3
        bg2 = 'FFF9C4' if yr in (2018, 2022) else ('F4F4F4' if ri % 2 == 0 else 'FFFFFF')
        for ci, v in enumerate([yr, f'{sh:+.1f}pp', f'{wt:.3f}',
                                 f'{sh*wt:+.3f}pp', cycle_notes.get(yr, '')], 1):
            cell = ws1.cell(row=r, column=ci, value=v)
            val(cell, bg=bg2, sz=9, ha='left' if ci == 5 else 'center')
        rh(ws1, r, 16)

    total_row = len(oop['cycle_weights']) + 3
    ws1.merge_cells(f'A{total_row}:E{total_row}')
    c = ws1.cell(row=total_row, column=1,
                 value=f"★ RESULT:  OOP shift = {oop['oop_shift']:+.2f}pp  |  "
                       f"σ(OOP) = ±{oop['oop_std']:.2f}pp  |  "
                       f"Combined σ = ±{oop['gcb_sigma']:.2f}pp  |  "
                       f"GCB eday = D{oop['gcb_eday_central']:+.2f}")
    hdr(c, bg='1B4F8A', sz=10, left=True)
    rh(ws1, total_row, 22)

    for i, w in enumerate([8, 12, 14, 14, 45], 1):
        cw(ws1, i, w)

    # ── SHEET 2: DEMOGRAPHIC SHIFTS ───────────────────────────────────────────
    ws2 = wb.create_sheet('2. Demographic Shifts')
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:G1')
    hdr(ws2.cell(row=1, column=1,
                 value='DEMOGRAPHIC SHIFTS vs CATALIST 2024  |  '
                       'Recency × grade weighted from manually entered crosstabs'),
        sz=11, left=True)
    rh(ws2, 1, 24)

    baselines = config['catalist_2024']['group_d_share']
    for ci, h in enumerate(['Group', 'Catalist 2024', 'Current Est.',
                             'Shift (pp)', 'N Entries', 'Last Updated', 'Note'], 1):
        hdr(ws2.cell(row=2, column=ci, value=h), bg='2C3E6B', sz=9)
    rh(ws2, 2, 20)

    for ri, (group, data) in enumerate(demo.items()):
        r = ri + 3
        bg2 = 'EBF3FB' if group in ('white_nc', 'hispanic', 'black') else \
              ('F4F4F4' if ri % 2 == 0 else 'FFFFFF')
        shift = data['shift']
        shift_bg = 'C8E6C9' if shift >= 2 else ('FFCDD2' if shift <= -2 else bg2)

        row_data = [
            group, baselines.get(group, ''),
            data['current_d_share'], shift,
            data['n_entries'], data.get('last_updated', ''),
            data.get('note', '')
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=ci, value=v)
            bg_use = shift_bg if ci == 4 else bg2
            val(cell, bg=bg_use, sz=9, ha='center' if ci < 7 else 'left')
            if ci in (2, 3) and isinstance(v, float):
                cell.number_format = '0.000'
            if ci == 4 and isinstance(v, float):
                cell.number_format = '+0.0;-0.0;0.0'
                cell.font = Font(name='Calibri', size=9, bold=True,
                                 color='1B4F8A' if v >= 0 else '8B1A1A')
        rh(ws2, r, 16)

    for i, w in enumerate([18, 13, 12, 11, 10, 13, 40], 1):
        cw(ws2, i, w)

    # ── SHEET 3: STATE ENVIRONMENTS ───────────────────────────────────────────
    ws3 = wb.create_sheet('3. State Environments')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A3'
    ws3.merge_cells('A1:H1')
    hdr(ws3.cell(row=1, column=1,
                 value=f'STATE GCB ELECTION DAY PROJECTIONS  |  '
                       f'Catalist 2024 + demographic shifts + OOP {oop["oop_shift"]:+.2f}pp'),
        sz=11, left=True)
    rh(ws3, 1, 24)

    for ci, h in enumerate(['Abbr', 'Catalist Baseline', 'Demo Shift',
                             'OOP Shift', 'GCB Eday', 'Low (−1σ)',
                             'High (+1σ)', 'Reliable'], 1):
        hdr(ws3.cell(row=2, column=ci, value=h), bg='2C3E6B', sz=9)
    rh(ws3, 2, 20)

    sorted_states = sorted(state_envs.items(),
                           key=lambda x: -(x[1].get('gcb_eday') or -99))
    for ri, (abbr, env) in enumerate(sorted_states):
        r = ri + 3
        gcb = env.get('gcb_eday')
        bg2 = 'F4F4F4' if ri % 2 == 0 else 'FFFFFF'

        row_data = [
            abbr,
            env.get('catalist_baseline'),
            env.get('demo_shift'),
            env.get('oop_shift'),
            gcb,
            env.get('gcb_eday_lo'),
            env.get('gcb_eday_hi'),
            'Yes' if env.get('reliable', True) else 'No',
        ]
        for ci, v in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=ci, value=v)
            val(cell, bg=bg2, sz=9)
            if ci in (2, 3, 4, 5, 6, 7) and isinstance(v, float):
                cell.number_format = '+0.00;-0.00;0.00'
            if ci == 5 and isinstance(v, (int, float)):
                fg = '1B4F8A' if (v or 0) >= 0 else '8B1A1A'
                cell.font = Font(name='Calibri', bold=True, size=9, color=fg)
        rh(ws3, r, 15)

    for i, w in enumerate([6, 15, 12, 10, 12, 12, 12, 9], 1):
        cw(ws3, i, w)

    # ── SAVE ──────────────────────────────────────────────────────────────────
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f'  Excel saved: {out_path}')
    return out_path


if __name__ == '__main__':
    from datetime import date
    import pandas as pd
    from model.oop_shift import compute_oop_shift
    from model.demographic_shifts import compute_group_shifts
    from model.state_environment import compute_state_environments
    from model.poll_blend import compute_poll_blend
    from model.race_model import compute_race, SENATE_RACES_2026
    from model.monte_carlo import run_simulation
    import yaml

    with open('config.yaml') as f:
        cfg = yaml.safe_load(f)

    oop  = compute_oop_shift(cfg)
    demo = compute_group_shifts(cfg)
    envs, _ = compute_state_environments(oop, demo, cfg)
    polls_df = pd.read_csv('data/state_polls.csv')
    ref = date.today()
    days = (date(2026, 11, 3) - ref).days

    race_results = {}
    for abbr, inc, dc, rc, comp in SENATE_RACES_2026:
        if abbr not in envs: continue
        blend = compute_poll_blend(abbr, polls_df, ref, cfg)
        race_results[abbr] = compute_race(abbr, envs[abbr], blend, days, cfg)

    sim = run_simulation(race_results, oop, cfg)
    ds = date.today().strftime('%Y-%m-%d')
    build_excel(oop, demo, envs, race_results, sim, cfg,
                f'outputs/senate_model_{ds}.xlsx')
    print('Done.')
